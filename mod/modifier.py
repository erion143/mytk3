import os
import sys
sys.path.append('..')
sys.path.append('.')


from tkinter import *
from tkinter.simpledialog import askstring
from mytk3 import TextWall
from graph3 import Deckart3, Line3, Point3, Modifier, BasicPoint3, bust
from myparser import parser
import openpyxl as xls
from xlstest import write_to_xls, create_or_rewrite


BEGIN = '1.0'


class Mod(Modifier):
    def __init__(self, master):
        Modifier.__init__(self, master)

    def fill(self, **kwargs):
        if not self.line:
            return None
        else:
            fp = self.line.points[0]
            lp = self.line.points[-1]
            start = Point3.create_from_point(fp,
                                             parent=fp,
                                             storage=self.points,
                                             **kwargs)
            print('!!!!!!!!!!!!!!1points', id(self.points), id(self.master.points))
            stop = Point3.create_from_point(lp,
                                            parent=lp,
                                            storage=self.points,
                                            **kwargs)
            print('points!!!!!!!!!!!!!', self.points)
            self.newline = Line3(self.master,
                                 self.points,
                                 color='#eea500',
                                 width=3,
                                 smooth=1)
            self.newline.points = self.points
            print(self.points)

            self.newline.replace()
            self.replace()

    def add_point(self, event, **kwargs):
        Point3.create_from_abs_coord(self.master,
                                     event.x, event.y,
                                     storage=self.points,
                                     **kwargs)

        self.points.sort(key=(lambda p: p.x))

        self.newline.replace()
        print('points2:', self.points)
        self.replace()
        print('points3:', self.points)

    def rm_point(self, event):
        inds = [p.ind for p in self.points]
        indr = self.master.canvas.find_closest(event.x, event.y)
        ind = bust(indr, inds)
        if ind != -1:
            p = self.points[ind]
        else:
            return None
        p.delete()
        self.newline.replace()
        self.replace()

    def save(self):
        print('\nLet\'s go save!\n')
        print('points:', self.points)
        self.line.points = []

        for p in self.points:
            self.line.points.append(BasicPoint3(p.master, p.x, p.y))

        self.line.replace()

    def run(self):
        Modifier.run(self)
        self.master.canvas.bind('<Double-Button-1>', self.add_point)
        self.master.canvas.bind('<Button-3>', self.rm_point)


class Container:
    def __init__(self):
        self.root = Tk()
        self.f_main = Frame(self.root)
        self.f_main.pack(fill=BOTH, expand=YES)  # temporary

        self.active = self.c_input = DataInput(self)
        self.active.grid()
        self.disable = self.c_plot = DataPlotter(self)

        self.root.bind('<Control-q>', self.switch)
        self.root.bind('<Control-a>', self.c_plot.size)
        self.root.bind('<Control-w>', self.c_input.f_go)
        self.root.bind('<Control-m>', self.c_plot.modify)
        self.root.bind('<Control-s>', self.c_plot.save)
        self.root.bind('<Control-l>', self.c_plot.show_level)

    def switch(self, event=None):
        print('switch')
        self.active.grid_forget()
        self.active, self.disable = self.disable, self.active
        self.active.grid()

    def mainloop(self):
        self.root.mainloop()


class DataInput:
    def __init__(self, master):
        self.master = master

        self.text = TextWall(self.master.f_main, width=50)
        self.b_go = Button(self.master.f_main,
                           text='Go',
                           command=self.f_go)

    def grid(self):
        self.text.grid(row=0, column=0, sticky=NSEW)
        self.b_go.grid(row=1, column=0, sticky=EW)

        self.master.f_main.grid_columnconfigure(0, weight=1)
        self.master.f_main.grid_rowconfigure(0, weight=1)

    def grid_forget(self):
        self.text.grid_forget()
        self.b_go.grid_forget()

    def size(self, event):
        print(self.text.winfo_width(), self.text.winfo_height())
        print(self.text.winfo_reqwidth(), self.text.winfo_reqheight())

    def get(self):
        return self.text.get(BEGIN, END)

    def f_go(self, event=None):
        raw_data = self.get().split('\n')
        print(raw_data)
        xs, ys = parser(raw_data, (float, float), source_is_file=False, replacement=((',', '.'),))
        self.master.c_plot.add_line(xs, ys)
        self.master.c_plot.plot.auto_scale()
        self.master.switch()


class DataPlotter:
    def __init__(self, master):
        self.master = master

        self.origin = None     # Тут хранится исходный график
        self.ancillary = None  # Тут сплайн, если он есть
        self.level = None      # Тут линия "у = число"
        self.begin = []        # тут будук координаты исходного origin

        self.canvas = Canvas(self.master.f_main, width=700, height=500)
        self.plot = Deckart3(self.canvas)
        self.plot.modifier = Mod(self.plot)
        self.plot.create_scale()

        self.button_frame = Frame(self.master.f_main)

        self.b_mod = Button(self.button_frame,
                            text='Modify',
                            command=self.modify)
        self.b_mod.pack(fill=X)
        self.b_save = Button(self.button_frame,
                             text='Save',
                             state=DISABLED,
                             command=self.save)
        self.b_save.pack(fill=X)
        self.b_saveto = Button(self.button_frame,
                               text='Save to file',
                               command=self.f_save)
        self.b_saveto.pack(fill=X)
        self.b_savetoxls = Button(self.button_frame,
                                  text='Save to excel',
                                  command=self.f_save_to_xls,
                                  state=DISABLED)
        self.b_savetoxls.pack(fill=X)
        self.e_level = Entry(self.button_frame)
        self.e_level.pack(side=BOTTOM, fill=X)
        self.b_level = Button(self.button_frame,
                              text='Set level',
                              command=self.show_level)
        self.b_level.pack(side=BOTTOM, fill=X)

    def grid(self):
        self.canvas.grid(row=0, column=0, sticky=NSEW)
        self.button_frame.grid(row=0, column=1, sticky=NSEW)

    def grid_forget(self):
        self.canvas.grid_forget()
        self.button_frame.grid_forget()

    def size(self, event=None):
        print(self.canvas.winfo_width(), self.canvas.winfo_height())
        print(self.canvas.winfo_reqwidth(), self.canvas.winfo_reqheight())

    def add_point(self, x, y):
        self.plot.add_point(x, y)

    def add_group_of_points(self, xs, ys):
        self.plot.add_group_of_points(xs, ys)

    def add_line(self, xs, ys):
        self.origin = Line3.from_coords(self.plot, xs, ys)
        self.origin.replace()

    def modify(self, event=None):
        if not self.begin:
            self.begin = [self.origin.get_xs(),
                          self.origin.get_ys()]
        self.b_mod.config(text='Cancel', command=self.f_cancel)
        self.b_save.config(state=NORMAL)
        self.b_savetoxls.config(state=NORMAL)
        self.b_saveto.config(state=DISABLED)
        self.plot.modifier.accept(self.origin)
        self.plot.modifier.fill()
        self.plot.modifier.run()

    def f_cancel(self):
        self.plot.modifier.stop()
        self.b_save.config(state=DISABLED)
        self.b_saveto.config(state=NORMAL)
        self.b_mod.config(text='Modify', command=self.modify)

    def save(self, event=None):
        self.plot.modifier.save()
        self.plot.modifier.stop()
        self.b_save.config(state=DISABLED)
        self.b_saveto.config(state=NORMAL)
        self.b_mod.config(text='Modify', command=self.modify)

    def show_level(self, event=None):
        try:
            num = int(self.e_level.get())
        except ValueError:
            print('No value')
            return None
        level = Line3.from_coords(self.plot,
                                  (self.plot.x_min, self.plot.x_max),
                                  (num, num))
        level.replace()
        self.level = level
        self.b_level.config(text='Cancel', command=self.f_level_del)
        self.e_level.config(state=DISABLED)

    def f_level_del(self):
        self.level.delete()
        self.b_level.config(text='Set level', command=self.show_level)
        self.e_level.config(state=NORMAL)
        self.e_level.delete(0, END)

    def f_save(self):
        with open('test_out.txt', 'a') as f:
            for x, y in zip(self.origin.get_xs(), self.origin.get_ys()):
                print('{}\t{}'.format(x, y), file=f)
            print(file=f)

    def f_save_to_xls(self):
        name = askstring(title='Open', prompt='Enter the file name:')
        if not name:
            return None

        if '.xlsx' not in name:
            name += '.xlsx'

        if os.path.exists(name):
            wb = xls.load_workbook(name)
        else:
            wb = xls.Workbook()

        sheets = wb.sheetnames
        if 'Sheet' in sheets:
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

        sheetname = create_or_rewrite(wb)
        if not sheetname:
            return None

        write_to_xls(wb,
                     sheetname,
                     *self.begin,
                     line=1,
                     start_column='A')
        write_to_xls(wb,
                     sheetname,
                     self.origin.get_xs(),
                     self.origin.get_ys(),
                     line=1,
                     start_column='D')
        wb.save(name)
        wb.close()


if __name__ == '__main__':
    c = Container()
    c.mainloop()
